"""
Export school data (with enrichment) from MongoDB to Excel (.xlsx).
Supports filtering, chunked export for large datasets, and formatted headers.
"""

import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import EXPORT_DIR, MAX_ROWS_PER_FILE
from db.mongo_client import mongo
from utils.logger import get_logger

logger = get_logger(__name__)


def _cleanup_old_exports(label: str):
    """Remove old export files for the same filter label to avoid duplicates."""
    # Match both exact (schools_label.xlsx) and legacy timestamped (schools_label_*.xlsx)
    patterns = [
        os.path.join(EXPORT_DIR, f"schools_{label}.xlsx"),
        os.path.join(EXPORT_DIR, f"schools_{label}_*.xlsx"),
        os.path.join(EXPORT_DIR, f"schools_{label}_part*.xlsx"),
    ]
    removed = 0
    for pattern in patterns:
        for old_file in glob.glob(pattern):
            try:
                os.remove(old_file)
                logger.info("Removed old export: %s", os.path.basename(old_file))
                removed += 1
            except OSError as e:
                logger.warning("Could not remove %s: %s", old_file, e)
    if removed:
        logger.info("Cleaned up %d old export file(s).", removed)


# Column definitions: (header_name, mongo_field, column_width)
COLUMNS = [
    ("No", None, 6),
    ("NPSN", "npsn", 14),
    ("Nama Sekolah", "nama", 45),
    ("Bentuk Pendidikan", "bentukPendidikan", 22),
    ("Status", "statusSatuanPendidikan", 12),
    ("Jenjang", "jenjangPendidikan", 24),
    ("Akreditasi", "detail_akreditasi", 14),
    ("Pembina", "pembina", 35),
    ("Provinsi", "namaProvinsi", 25),
    ("Kabupaten/Kota", "namaKabupaten", 25),
    ("Kecamatan", "namaKecamatan", 22),
    ("Desa/Kelurahan", "namaDesa", 20),
    ("Alamat", "alamatJalan", 50),
    ("Kode Pos", "detail_kode_pos", 10),
    ("Kepala Sekolah", "detail_kepala_sekolah", 30),
    ("Telepon (Pemerintah)", "detail_phone", 22),
    ("Email (Pemerintah)", "detail_email", 30),
    ("Website (Pemerintah)", "detail_website", 40),
    ("Instagram", "instagram", 30),
    ("WhatsApp", "whatsapp", 22),
    ("Facebook", "facebook", 35),
    ("Website (Sosmed)", "website_social", 40),
    ("Nomor Kontak (Sosmed)", "contact_social", 22),
    ("Email (Sosmed)", "email", 30),
    ("Jumlah Siswa", "detail_jumlah_siswa", 14),
    ("Siswa L", "detail_siswa_laki", 10),
    ("Siswa P", "detail_siswa_perempuan", 10),
    ("Jumlah Rombel", "detail_jumlah_rombel", 14),
    ("Guru L", "detail_guru_laki", 10),
    ("Guru P", "detail_guru_perempuan", 10),
    ("Kurikulum", "detail_kurikulum", 30),
    ("Yayasan", "detail_yayasan", 35),
    ("Latitude", "detail_lintang", 14),
    ("Longitude", "detail_bujur", 14),
]


def _style_header(ws):
    """Apply styling to the header row."""
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"


def _write_rows(ws, schools: list[dict], start_num: int = 1):
    """Write school data rows to the worksheet."""
    data_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_idx, school in enumerate(schools, 2):
        for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
            if field is None:
                # Row number
                value = start_num + row_idx - 2
            else:
                value = school.get(field, "") or ""

            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.alignment = data_alignment
            cell.border = thin_border

    return len(schools)


def export_to_excel(
    output_path: str | None = None,
    filters: dict | None = None,
    province_name: str = "",
):
    """
    Export school data with enrichment to Excel.

    Args:
        output_path: Full path for the output .xlsx file (auto-generated if None)
        filters: MongoDB query filters to select which schools to export
        province_name: Optional province name for filename
    """
    filters = filters or {}

    logger.info("=" * 60)
    logger.info("EXPORT: Schools to Excel")
    logger.info("=" * 60)

    # Ensure export directory exists
    os.makedirs(EXPORT_DIR, exist_ok=True)

    # Count total records
    total = mongo.get_school_count(filters)
    enriched = mongo.get_enriched_count(filters)
    logger.info("Schools matching filter: %d (enriched: %d)", total, enriched)

    if total == 0:
        logger.warning("No schools to export.")
        return None

    # Generate deterministic filename (no timestamp → overwrites previous export)
    label = province_name.replace(" ", "_").replace(".", "").lower() if province_name else "all"
    if not output_path:
        output_path = os.path.join(EXPORT_DIR, f"schools_{label}.xlsx")

    # Clean up old export files matching this filter to avoid duplicates
    _cleanup_old_exports(label)

    # Check if we need chunked export
    if total > MAX_ROWS_PER_FILE:
        return _export_chunked(output_path, filters, total)

    # Single-file export
    return _export_single(output_path, filters, total)


def _export_single(output_path: str, filters: dict, total: int) -> str:
    """Export all data to a single Excel file."""
    logger.info("Exporting %d records to: %s", total, output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "School Data"
    _style_header(ws)

    # Fetch and write in batches to avoid memory issues
    batch_size = 5000
    offset = 0
    row_num = 1

    while offset < total:
        schools = mongo.get_schools_with_enrichment(
            query=filters,
            skip=offset,
            limit=batch_size,
        )

        if not schools:
            break

        _write_rows(ws, schools, start_num=row_num)
        row_num += len(schools)
        offset += len(schools)
        logger.info("Written %d/%d rows...", offset, total)

    # Add summary sheet
    _add_summary_sheet(wb, filters, total)

    wb.save(output_path)
    logger.info("Export complete: %s (%d rows)", output_path, total)
    return output_path


def _export_chunked(base_path: str, filters: dict, total: int) -> list[str]:
    """Export data split across multiple Excel files."""
    logger.info(
        "Large dataset (%d records). Splitting into files of %d rows each.",
        total,
        MAX_ROWS_PER_FILE,
    )

    files = []
    offset = 0
    file_num = 1

    while offset < total:
        # Generate chunk filename
        base, ext = os.path.splitext(base_path)
        chunk_path = f"{base}_part{file_num}{ext}"

        wb = Workbook()
        ws = wb.active
        ws.title = f"School Data (Part {file_num})"
        _style_header(ws)

        chunk_written = 0
        row_num = offset + 1

        while chunk_written < MAX_ROWS_PER_FILE and (offset + chunk_written) < total:
            batch_size = min(5000, MAX_ROWS_PER_FILE - chunk_written)
            schools = mongo.get_schools_with_enrichment(
                query=filters,
                skip=offset + chunk_written,
                limit=batch_size,
            )

            if not schools:
                break

            _write_rows(ws, schools, start_num=row_num)
            row_num += len(schools)
            chunk_written += len(schools)

        wb.save(chunk_path)
        files.append(chunk_path)
        logger.info(
            "Part %d: %s (%d rows)",
            file_num,
            chunk_path,
            chunk_written,
        )

        offset += chunk_written
        file_num += 1

    logger.info("Chunked export complete: %d files", len(files))
    return files


def _add_summary_sheet(wb: Workbook, filters: dict, total: int):
    """Add a summary/metadata sheet to the workbook."""
    ws = wb.create_sheet("Summary")

    header_font = Font(name="Calibri", bold=True, size=12)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    rows = [
        ("Export Information", ""),
        ("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Records", str(total)),
        ("Filters Applied", str(filters) if filters else "None"),
        ("Source", "Kemendikdasmen (data.belajar.id + detail API)"),
        ("Enrichment", "Google Search + DeepSeek AI"),
        ("", ""),
        ("Column Descriptions", ""),
        ("NPSN", "Nomor Pokok Sekolah Nasional (unique school ID)"),
        ("Telepon (Pemerintah)", "Phone number from government detail API"),
        ("Email (Pemerintah)", "Email from government detail API"),
        ("Instagram", "School Instagram account (from Google/DeepSeek)"),
        ("WhatsApp", "School WhatsApp number (from Google/DeepSeek or govt)"),
        ("Facebook", "School Facebook page (from Google/DeepSeek)"),
        ("Website (Sosmed)", "School website found via social media search"),
        ("Nomor Kontak (Sosmed)", "Contact number from social media search"),
        ("Kepala Sekolah", "Principal name from government data"),
    ]

    for row_idx, (key, value) in enumerate(rows, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=key)
        cell_b = ws.cell(row=row_idx, column=2, value=value)
        if row_idx in (1, 8):
            cell_a.font = header_font
